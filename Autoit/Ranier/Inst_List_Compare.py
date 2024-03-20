import openpyxl
from openpyxl.utils import get_column_letter

# Define the file paths
file_path1 = "C:/Users/rgrov/OneDrive - Bit-Wise Automation LLC/Richards_Docs/My_Scripts/Ranier/22133-JRFML Project Rainier-Instruments List 12-15-23.xls"
file_path2 = "C:/Users/rgrov/OneDrive - Bit-Wise Automation LLC/Richards_Docs/My_Scripts/Ranier/22133-JRFML Project Rainier-Instrument List 2-7-24.xls"

# Function to open an Excel workbook
def open_workbook(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        return workbook, workbook.active
    except Exception as e:
        print(f"Failed to open {file_path}: {e}")
        return None, None

# Open the workbooks
workbook1, sheet1 = open_workbook(file_path1)
workbook2, sheet2 = open_workbook(file_path2)

if not sheet1 or not sheet2:
    raise Exception("One of the workbooks could not be opened")

# Function to find the row index for a given PnPGuid
def find_pnpguid_index(sheet, pnpguid):
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers are in the first row
        if row[8] == pnpguid:  # Assuming PnPGuid is in the 9th column (column I)
            return row
    return None

# Prepare the headers for the new workbook
headers = ["New Tag", "Old Tag", "New Description", "Old Description", "New Connection Type", "Old Connection Type",
           "New Connection Size", "Old Connection Size", "New DWG Number", "Old DWG Number", "New Supplier", "Old Supplier",
           "New Comment", "Old Comment", "New PnPID", "Old PnPID", "PnPGuid", "New Device", "Deleted Device"]

# Initialize a new workbook for output
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.append(headers)

# Compare the data and fill in the new workbook
# Helper function to get row data as a dictionary
def get_row_data(row):
    return {f"column_{i+1}": cell.value for i, cell in enumerate(row)}

# Helper function to find a PnPGuid in a sheet and return its row as a dictionary
def find_pnpguid(sheet, pnpguid):
    for row in sheet.iter_rows(min_row=2):
        if row[8].value == pnpguid:  # Assuming PnPGuid is in the 9th column (index 8)
            return get_row_data(row)
    return None

# Iterate through the first sheet to find changes or deletions
for row1 in sheet1.iter_rows(min_row=2, values_only=True):
    pnpguid = row1[8]  # PnPGuid column
    row_data_sheet2 = find_pnpguid(sheet2, pnpguid)
    
    if row_data_sheet2:
        # Compare each column data except PnPGuid itself
        changes = [pnpguid]
        for i, (value1, value2) in enumerate(zip(row1, row_data_sheet2.values())):
            if i != 8:  # Skip PnPGuid column for comparison
                if value1 != value2:
                    changes.extend([value2, value1])  # New value from sheet2, old value from sheet1
                else:
                    changes.extend(["", ""])  # No change in this column
        new_sheet.append(changes)
    else:
        # PnPGuid deleted
        deleted_row = [""] * 17 + [pnpguid, "x"]  # Mark as deleted
        new_sheet.append(deleted_row)

# Check for new PnPGuids in the second sheet
for row2 in sheet2.iter_rows(min_row=2, values_only=True):
    pnpguid = row2[8]  # PnPGuid column
    if not find_pnpguid(sheet1, pnpguid):
        # New PnPGuid found
        new_row = [""] * 17 + [pnpguid, "", "x"]  # Mark as new
        new_sheet.append(new_row)

# Save the new workbook
output_file_path = "comparison_result.xlsx"
new_workbook.save(output_file_path)

# Close the opened workbooks
workbook1.close()
workbook2.close()

print(f"Comparison completed. Results saved in {output_file_path}")
